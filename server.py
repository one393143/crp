"""
server.py — CRP 戰情系統 Web 介面
------------------------------------
使用方式：python server.py
瀏覽器會自動開啟 http://localhost:5000
"""
import sys, io, json, webbrowser, glob, re
from pathlib import Path
from datetime import datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

try:
    from flask import Flask, jsonify, send_from_directory, request
except ImportError:
    print("[錯誤] 請先安裝 Flask：python -m pip install flask")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("[錯誤] 請先安裝 openpyxl：python -m pip install openpyxl")
    sys.exit(1)

import db

app = Flask(__name__, static_folder=str(Path(__file__).parent))

# ── 系統欄位（永遠略過）──────────────────────────────
SKIP = {"$id", "$revision", "記錄號碼", "建立人", "更新人",
        "建立時間", "更新時間", "執行者", "類別", "狀態"}

# ── App 71 聯絡人 Field Code → Label（順序即為欄位順序）──
SCHEMA_71 = {
    "廠商代號":            "客戶/供應商代號",
    "類型":               "客戶/供應商類型",
    "客戶_供應商名稱":      "客戶/供應商名稱",
    "客戶_廠商聯絡人姓名":  "聯絡人姓名",
    "職稱":               "聯絡人職稱",
    "聯絡人電話":          "聯絡人電話",
    "分機":               "聯絡人分機",
    "手機":               "聯絡人手機",
    "聯絡人Email":        "聯絡人Email",
    "聯絡人地址":          "聯絡人地址",
}

# ── App 69 訂購單 ───────────────────────────────────────
# Header 欄位（訂購單層級）
SCHEMA_69_HEADER = {
    "訂購單號":     "訂購單編號",
    "狀態":         "狀態",
    "出貨狀態":     "出貨狀態",
    "客戶名稱":     "客戶名稱",
    "訂購日期":     "訂購日期",
    "客戶訂單號碼": "客戶訂單號",
    "負責業務":     "負責業務",
    "幣別":         "客戶訂單幣別",
    "金額合計":     "訂單金額總計",
    "本幣金額合計": "訂單金額總計(TWD)",
}
# Detail 欄位（訂購明細 subtable 內）
SCHEMA_69_DETAIL = {
    "行號":      "行號",
    "產品名稱":   "產品名稱",
    "產品代號":   "EIT產品編號",
    "供應商代號":   "供應商代號",
    "供應商名稱":   "供應商名稱",
    "數量":      "數量",
    "單價":      "單價",
    "金額":      "金額合計",
    "金額合計_TWD_": "金額合計(TWD)",
    "預計交期":   "客戶預計交期",
    "日期":      "EIT進貨日期",
    "幣別日期":   "幣別日期",
    "成本匯率":   "成本匯率",
    "成本幣別":   "成本幣別",
    "成本單價":   "成本單價",
    "成本金額_TWD_": "成本金額(TWD)",
    "利潤率":   "利潤率",
    "採購單號":   "採購單編號",
    "已轉採購量":  "已轉採購量",
    "未轉採購量":  "未轉採購量",
    "銷貨數量":   "銷貨數量",
    "未銷貨數量":  "未銷貨數量",
}
# 展平的完整欄位順序
SCHEMA_69_COLS = list(SCHEMA_69_HEADER.values()) + list(SCHEMA_69_DETAIL.values())

# ── App 310 POS 歷史 ─────────────────────────────────
SCHEMA_310 = {
    "單行文字方塊_2":             "Customer Name",
    "日期":                       "PO Date",
    "Shipping_Date":             "Shipping Date",
    "Payment_Date":              "Payment Date",
    "單行文字方塊_15":            "Due Date",
    "單行文字方塊_8":             "Product Type",
    "單行文字方塊_9":             "Product No.",
    "單行文字方塊_1":             "Type",
    "Qty_":                      "Qty.",
    "Sales_Unit_Price__NT__":    "Sales Unit Price (NT$)",
    "Sales_Total_Price___NT__":  "Sales Total (NT$)",
    "Cost_Unit_Price___NT__":    "Cost Unit Price (NT$)",
    "Cost_Total_Price__NT__":    "Cost Total (NT$)",
    "單行文字方塊_19":            "Margin",
    "單行文字方塊_4":             "EIT PO No.",
    "單行文字方塊_3":             "Customer PO No.",
    "單行文字方塊_5":             "PI(PL) No.",
    "單行文字方塊_6":             "Invoice No.",
    "單行文字方塊_0":             "Sales",
    "單行文字方塊_16":            "Country",
    "數值_1":                    "Exchange Rate",
    "單行文字方塊_18":            "Remark",
    "單行文字方塊":               "No.",
}


def _val(v):
    """從 Kintone field value 提取純文字"""
    if v is None:
        return ""
    if isinstance(v, list):
        if not v:
            return ""
        if isinstance(v[0], dict) and "name" in v[0]:
            return ", ".join(i.get("name", "") for i in v if i.get("name"))
        return ", ".join(str(i) for i in v if str(i).strip())
    if isinstance(v, dict):
        return str(v.get("name", "")).strip()
    return str(v).strip()


def to_row(raw_json_str: str, schema: dict) -> list:
    """將 raw_json 轉為按 schema 欄位順序的 list"""
    cols = list(schema.values())
    try:
        rec = json.loads(raw_json_str)
    except Exception:
        return [""] * len(cols)

    labeled = {}
    for code, label in schema.items():
        field = rec.get(code)
        if not field or not isinstance(field, dict):
            continue
        if field.get("type") == "SUBTABLE":
            continue
        labeled[label] = _val(field.get("value"))

    return [labeled.get(col, "") for col in cols]


def _fmt_num(val):
    if val is None:
        return ""
    try:
        f = float(val)
        if f == int(f):
            return str(int(f))
        return f"{f:.2f}"
    except (ValueError, TypeError):
        return str(val)


def orders_to_rows(raw_json_str: str) -> list[list]:
    """展平訂購單：每個明細行輸出一列，前綴帶訂單 header 欄位，自動補齊與計算缺失欄位"""
    try:
        rec = json.loads(raw_json_str)
    except Exception:
        return []

    # 1. 取得貨幣與匯率資訊
    currency = _val(rec.get("幣別", {}).get("value", ""))
    rate_val = rec.get("匯率", {}).get("value")
    ex_rate = _safe_float(rate_val, default=None)
    
    hdr_total_raw = rec.get("金額合計", {}).get("value")
    hdr_twd_raw = rec.get("本幣金額合計", {}).get("value")
    
    if ex_rate is None or ex_rate == 0.0:
        t = _safe_float(hdr_total_raw)
        twd = _safe_float(hdr_twd_raw)
        if t and twd:
            ex_rate = twd / t
        else:
            ex_rate = 1.0

    # 2. 處理明細行，計算明細總和
    detail_table = rec.get('訂購明細', {}).get('value', [])
    processed_details = []
    detail_amt_sum = 0.0
    detail_twd_sum = 0.0
    
    for line_item in detail_table:
        cell = line_item.get('value', {})
        qty = _safe_float(cell.get("數量", {}).get("value"))
        price = _safe_float(cell.get("單價", {}).get("value"))
        
        # 金額合計 (detail)
        cell_amt_raw = cell.get("金額", {}).get("value")
        cell_amt = _safe_float(cell_amt_raw, default=None)
        if cell_amt is None or cell_amt == 0.0:
            if qty and price:
                cell_amt = qty * price
            else:
                cell_amt = 0.0
        
        # 金額合計(TWD) (detail)
        cell_twd_raw = cell.get("金額合計_TWD_", {}).get("value")
        cell_twd = _safe_float(cell_twd_raw, default=None)
        if cell_twd is None or cell_twd == 0.0:
            cell_twd = cell_amt * ex_rate
            
        detail_amt_sum += cell_amt
        detail_twd_sum += cell_twd
        
        # 建立此行的 detail 欄位對照
        detail_vals = {}
        for code, label in SCHEMA_69_DETAIL.items():
            if code == "金額":
                detail_vals[label] = _val(cell_amt_raw) if cell_amt_raw else _fmt_num(cell_amt)
            elif code == "金額合計_TWD_":
                detail_vals[label] = _val(cell_twd_raw) if cell_twd_raw else _fmt_num(cell_twd)
            elif code == "利潤率":
                val = _val(cell.get(code, {}).get('value')) if cell.get(code) else ''
                if val and not val.endswith('%'):
                    val = f"{val}%"
                detail_vals[label] = val
            else:
                fv = cell.get(code)
                detail_vals[label] = _val(fv.get('value')) if fv else ''
        processed_details.append(detail_vals)

    # 3. 處理 Header 欄位（若欄位為空則用明細總和補齊）
    hdr_total = _safe_float(hdr_total_raw, default=None)
    if hdr_total is None or hdr_total == 0.0:
        hdr_total = detail_amt_sum
        
    hdr_twd = _safe_float(hdr_twd_raw, default=None)
    if hdr_twd is None or hdr_twd == 0.0:
        hdr_twd = detail_twd_sum

    header_vals = []
    for code, label in SCHEMA_69_HEADER.items():
        if code == "金額合計":
            header_vals.append(_val(hdr_total_raw) if hdr_total_raw else _fmt_num(hdr_total))
        elif code == "本幣金額合計":
            header_vals.append(_val(hdr_twd_raw) if hdr_twd_raw else _fmt_num(hdr_twd))
        else:
            field = rec.get(code)
            if field and isinstance(field, dict) and field.get('type') != 'SUBTABLE':
                header_vals.append(_val(field.get('value')))
            else:
                header_vals.append('')

    # 4. 合併輸出
    if not processed_details:
        return [header_vals + [''] * len(SCHEMA_69_DETAIL)]
        
    result = []
    for detail_vals in processed_details:
        line_row = []
        for label in SCHEMA_69_DETAIL.values():
            line_row.append(detail_vals.get(label, ''))
        result.append(header_vals + line_row)
    return result



RECORD_DIR = Path(__file__).parent / "record"


def _find_latest(pattern: str) -> Path | None:
    """用 glob pattern 找最新（檔名日期最大）的檔案"""
    files = glob.glob(str(RECORD_DIR / pattern))
    if not files:
        return None
    # 從檔名中抽出數字串，取最大值
    def key(f):
        nums = re.findall(r'\d+', Path(f).stem)
        return nums[-1] if nums else ''
    return Path(sorted(files, key=key)[-1])


def _fmt_date(v) -> str:
    """將 datetime 或字串統一轉為 YYYY-MM-DD"""
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if v is None:
        return ''
    s = str(v).strip()
    # 常見格式：2025/12/04
    s = s.replace('/', '-')
    return s[:10] if len(s) >= 10 else s


def read_eit_logs() -> dict:
    """讀取 EIT Sales Report 接洽紀錄 Sheet（支援「接洽紀錄」與「接洽紀錄.」），回傳完整欄位"""
    COLS = ['聯絡時間', '接洽類型', '客戶名稱', '聯絡人', '接洽方式', '接洽情況', '接洽紀錄', '主管評價']
    rows = []
    seen = set()
    path = _find_latest('EIT Sales Report Chauncey*.xlsx')
    if not path:
        return {'columns': COLS, 'rows': [], 'error': '找不到 EIT Sales Report 檔案'}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        target_sheets = [n for n in wb.sheetnames if '接洽紀錄' in n]
        for sh in target_sheets:
            ws = wb[sh]
            col_map = {}
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    row_strs = [str(c).strip() if c else '' for c in row]
                    if '聯絡時間' in row_strs:
                        header_found = True
                        for ci, cell in enumerate(row):
                            label = str(cell).strip() if cell else ''
                            if label in COLS:
                                col_map[label] = ci
                    continue
                if not col_map:
                    continue
                date_val = row[col_map['聯絡時間']] if '聯絡時間' in col_map else None
                # 必須是 datetime 才是有效紀錄（過濾統計彙整列）
                if not isinstance(date_val, datetime):
                    continue
                date_str = _fmt_date(date_val)
                customer = str(row[col_map.get('客戶名稱', 2)] or '').strip()
                contact  = str(row[col_map.get('聯絡人', 3)] or '').strip()
                key = (date_str, customer, contact)
                if key in seen:
                    continue
                seen.add(key)
                rows.append([
                    date_str,
                    str(row[col_map.get('接洽類型', 1)] or '').strip(),
                    customer,
                    contact,
                    str(row[col_map.get('接洽方式', 4)] or '').strip(),
                    str(row[col_map.get('接洽情況', 5)] or '').strip(),
                    str(row[col_map.get('接洽紀錄', 6)] or '').strip(),
                    str(row[col_map.get('主管評價', 7)] or '').strip(),
                ])
        wb.close()
    except Exception as e:
        return {'columns': COLS, 'rows': [], 'error': str(e)}
    rows.sort(key=lambda r: r[0] or '', reverse=True)
    return {'columns': COLS, 'rows': rows}


def read_tw_opp() -> dict:
    """讀取 TW客戶狀態紀錄 New Business Opp Sheet，動態建立完整欄位"""
    path = _find_latest('TW客戶狀態紀錄*.xlsx')
    if not path:
        return {'columns': [], 'rows': [], 'error': '找不到 TW客戶狀態紀錄 檔案'}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        opp_sheet = next((n for n in wb.sheetnames if 'New Business' in n), None)
        if not opp_sheet:
            wb.close()
            return {'columns': [], 'rows': [], 'error': '找不到 New Business Opp 工作表'}
        ws = wb[opp_sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return {'columns': [], 'rows': [], 'error': str(e)}

    if not all_rows:
        return {'columns': [], 'rows': []}

    # 第一行是 header，跳過 None 欄（如序號欄）
    header_row = all_rows[0]
    columns = []
    col_indices = []
    for ci, cell in enumerate(header_row):
        if cell and str(cell).strip():
            columns.append(str(cell).strip())
            col_indices.append(ci)

    # 找「公司」欄索引用於過濾空白行
    company_col_pos = next((i for i, c in enumerate(columns) if '公司' in c), None)

    result_rows = []
    for row in all_rows[1:]:
        if not any(row):
            continue
        if company_col_pos is not None:
            ci = col_indices[company_col_pos]
            val = row[ci] if len(row) > ci else None
            if not val or not str(val).strip():
                continue
        out = []
        for ci in col_indices:
            v = row[ci] if len(row) > ci else None
            if isinstance(v, datetime):
                out.append(_fmt_date(v))
            elif v is None:
                out.append('')
            else:
                out.append(str(v).strip())
        result_rows.append(out)

    return {'columns': columns, 'rows': result_rows}


# ── Routes ────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(str(Path(__file__).parent), "index.html")


@app.route("/api/status")
def api_status():
    return jsonify({
        "sync_logs": db.get_sync_status(),
        "counts":    db.get_db_counts(),
    })


@app.route("/api/contacts")
def api_contacts():
    rows = db.search_contacts("", limit=5000)
    cols = list(SCHEMA_71.values())
    return jsonify({
        "columns": cols,
        "rows": [to_row(r["raw_json"], SCHEMA_71) for r in rows],
    })


@app.route("/api/orders")
def api_orders():
    db_rows = db.search_orders("", limit=5000)
    result_rows = []
    for r in db_rows:
        result_rows.extend(orders_to_rows(r["raw_json"]))
    return jsonify({
        "columns": SCHEMA_69_COLS,
        "rows":    result_rows,
    })


@app.route("/api/pos")
def api_pos():
    rows = db.search_pos("", limit=20000)
    cols = list(SCHEMA_310.values())
    return jsonify({
        "columns": cols,
        "rows": [to_row(r["raw_json"], SCHEMA_310) for r in rows],
    })


@app.route("/api/eit_logs")
def api_eit_logs():
    return jsonify(read_eit_logs())


@app.route("/api/tw_opp")
def api_tw_opp():
    return jsonify(read_tw_opp())


# ── Analysis APIs ─────────────────────────────────────

def _safe_float(v, default=0.0):
    try:
        return float(v) if v not in (None, '', 'None') else default
    except (ValueError, TypeError):
        return default

def _year(date_str):
    """從日期字串取年份"""
    if not date_str:
        return None
    s = str(date_str).strip()
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return None


def get_parsed_transactions(order_records, pos_records, q=None, filter_type=None):
    import json
    
    def parse_d(d_str):
        if not d_str: return None
        try:
            return datetime.strptime(d_str.strip()[:10], "%Y-%m-%d")
        except:
            return None
            
    # 1. Parse Order Detail Lines
    order_lines = []
    seen_lines = set()
    for r in order_records:
        try:
            rec = json.loads(r["raw_json"])
        except Exception:
            continue
        order_no = _val(rec.get("訂購單號", {}).get("value", ""))
        order_date_str = _val(rec.get("訂購日期", {}).get("value", ""))
        year = _year(order_date_str)
        if not year:
            continue
            
        rate_val = rec.get("匯率", {}).get("value")
        ex_rate = _safe_float(rate_val, default=None)
        
        hdr_total_raw = rec.get("金額合計", {}).get("value")
        hdr_twd_raw = rec.get("本幣金額合計", {}).get("value")
        
        if ex_rate is None or ex_rate == 0.0:
            t = _safe_float(hdr_total_raw)
            twd_hdr = _safe_float(hdr_twd_raw)
            if t and twd_hdr:
                ex_rate = twd_hdr / t
            else:
                ex_rate = 1.0

        detail_table = rec.get("訂購明細", {}).get("value", [])
        for line_item in detail_table:
            cell = line_item.get("value", {})
            line_no = _val(cell.get("行號", {}).get("value", ""))
            
            key = (order_no, line_no)
            if key in seen_lines:
                continue
            seen_lines.add(key)
            
            product_no = _val(cell.get("產品代號", {}).get("value", "")).strip()
            product_name = _val(cell.get("產品名稱", {}).get("value", "")).strip() # category
            
            # Apply filters
            if filter_type == 'product' and q:
                if q.lower() not in product_no.lower():
                    continue
            elif filter_type == 'category' and q:
                if q.lower() not in product_name.lower():
                    continue
                    
            qty = _safe_float(cell.get("數量", {}).get("value"))
            price = _safe_float(cell.get("單價", {}).get("value"))
            
            cell_amt = _safe_float(cell.get("金額", {}).get("value"), default=None)
            if cell_amt is None or cell_amt == 0.0:
                cell_amt = qty * price
                
            twd = _safe_float(cell.get("金額合計_TWD_", {}).get("value"), default=None)
            if twd is None or twd == 0.0:
                twd = cell_amt * ex_rate
                
            cost = _safe_float(cell.get("成本金額_TWD_", {}).get("value"), default=None)
            if cost is None or cost == 0.0:
                cost_price = _safe_float(cell.get("成本單價", {}).get("value"))
                cost_rate = _safe_float(cell.get("成本匯率", {}).get("value"), default=ex_rate)
                cost = qty * cost_price * cost_rate
                
            customer = _val(rec.get("客戶名稱", {}).get("value", ""))
            
            order_lines.append({
                'source': 'orders',
                'order_no': order_no,
                'customer': customer,
                'date_str': order_date_str,
                'date': parse_d(order_date_str),
                'year': year,
                'product_no': product_no,
                'category': product_name,
                'qty': qty,
                'amount_twd': twd,
                'cost_twd': cost,
                'matched': False
            })

    # 2. Parse POS Lines
    pos_lines = []
    for r in pos_records:
        po_date_str = r.get("po_date", "")
        year = _year(po_date_str)
        if not year:
            continue
            
        product_no = (r.get("product_no") or "").strip()
        product_name = (r.get("product_type") or "").strip() # category
        
        # Apply filters
        if filter_type == 'product' and q:
            if q.lower() not in product_no.lower():
                continue
        elif filter_type == 'category' and q:
            if q.lower() not in product_name.lower():
                continue
                
        qty = _safe_float(r.get("qty"))
        sales_total = _safe_float(r.get("sales_total"))
        
        try:
            p_raw = json.loads(r["raw_json"])
            cost_twd = _safe_float(p_raw.get("Cost_Total_Price__NT__", {}).get("value"))
        except:
            cost_twd = 0.0
            
        customer = r.get("customer", "")
        
        pos_lines.append({
            'source': 'pos',
            'customer': customer,
            'date_str': po_date_str,
            'date': parse_d(po_date_str),
            'year': year,
            'product_no': product_no,
            'category': product_name,
            'qty': qty,
            'amount_twd': sales_total,
            'cost_twd': cost_twd,
            'matched': False
        })
        
    # 3. Match 1-to-1
    for o in order_lines:
        if not o['date']: continue
        for p in pos_lines:
            if p['matched']: continue
            if not p['date']: continue
            
            # Match condition: same product, same qty, date within 7 days
            if o['product_no'].lower() != p['product_no'].lower():
                continue
            if abs(o['qty'] - p['qty']) > 0.01:
                continue
            if abs((o['date'] - p['date']).days) <= 7:
                o['matched'] = True
                p['matched'] = True
                break
                
    # Combine POS + unmatched Orders
    merged = []
    for p in pos_lines:
        merged.append(p)
    for o in order_lines:
        if not o['matched']:
            merged.append(o)
            
    return merged


@app.route("/api/analysis/customer")
def api_analysis_customer():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "請提供搜尋關鍵字"}), 400

    order_records = db.search_orders(q, limit=5000)
    pos_records = db.search_pos(q, limit=10000)
    merged = get_parsed_transactions(order_records, pos_records, filter_type=None)

    yearly_t = {}
    yearly_p = {}
    yearly_c = {}
    
    for tx in merged:
        yr = tx['year']
        source = tx['source']
        amt = tx['amount_twd']
        qty = tx['qty']
        prod = tx['product_no']
        cat = tx['category']
        
        # 1. Totals
        if yr not in yearly_t:
            yearly_t[yr] = {'pos_sum': 0.0, 'orders_sum': 0.0}
        if source == 'pos':
            yearly_t[yr]['pos_sum'] += amt
        else:
            yearly_t[yr]['orders_sum'] += amt
            
        # 2. Products
        if prod:
            pk = (yr, prod)
            if pk not in yearly_p:
                yearly_p[pk] = {'pos_qty': 0.0, 'orders_qty': 0.0}
            if source == 'pos':
                yearly_p[pk]['pos_qty'] += qty
            else:
                yearly_p[pk]['orders_qty'] += qty
                
        # 3. Categories
        if cat:
            ck = (yr, cat)
            if ck not in yearly_c:
                yearly_c[ck] = {'pos_qty': 0.0, 'orders_qty': 0.0}
            if source == 'pos':
                yearly_c[ck]['pos_qty'] += qty
            else:
                yearly_c[ck]['orders_qty'] += qty

    # Format output
    yearly_totals = []
    for yr in sorted(yearly_t.keys()):
        p_sum = yearly_t[yr]['pos_sum']
        o_sum = yearly_t[yr]['orders_sum']
        tot = p_sum + o_sum
        if p_sum > 0 and o_sum > 0:
            src = 'orders & pos'
        elif p_sum > 0:
            src = 'pos'
        else:
            src = 'orders'
        yearly_totals.append({
            'year': yr,
            'amount_twd': tot,
            'source': src
        })
        
    yearly_products = []
    for (yr, prod), q_dict in sorted(yearly_p.items()):
        p_qty = q_dict['pos_qty']
        o_qty = q_dict['orders_qty']
        tot = p_qty + o_qty
        if tot > 0:
            if p_qty > 0 and o_qty > 0:
                src = 'orders & pos'
            elif p_qty > 0:
                src = 'pos'
            else:
                src = 'orders'
            yearly_products.append({
                'year': yr,
                'product': prod,
                'qty': tot,
                'source': src
            })
            
    yearly_categories = []
    for (yr, cat), q_dict in sorted(yearly_c.items()):
        p_qty = q_dict['pos_qty']
        o_qty = q_dict['orders_qty']
        tot = p_qty + o_qty
        if tot > 0:
            if p_qty > 0 and o_qty > 0:
                src = 'orders & pos'
            elif p_qty > 0:
                src = 'pos'
            else:
                src = 'orders'
            yearly_categories.append({
                'year': yr,
                'category': cat,
                'qty': tot,
                'source': src
            })

    return jsonify({
        "query": q,
        "yearly_totals": yearly_totals,
        "yearly_products": yearly_products,
        "yearly_categories": yearly_categories,
    })


@app.route("/api/analysis/product")
def api_analysis_product():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "請提供搜尋關鍵字"}), 400

    all_orders = db.search_orders("", limit=5000)
    all_pos = db.search_pos("", limit=20000)
    merged = get_parsed_transactions(all_orders, all_pos, q=q, filter_type='product')

    yearly_vol = {}
    yearly_rev_cost = {}
    cust_breakdown = {}
    
    for tx in merged:
        yr = tx['year']
        source = tx['source']
        qty = tx['qty']
        amt = tx['amount_twd']
        cost = tx['cost_twd']
        cust = tx['customer']
        
        # 1. Volume
        if yr not in yearly_vol:
            yearly_vol[yr] = {'pos_qty': 0.0, 'orders_qty': 0.0}
        if source == 'pos':
            yearly_vol[yr]['pos_qty'] += qty
        else:
            yearly_vol[yr]['orders_qty'] += qty
            
        # 2. Revenue & Cost (for margin)
        if yr not in yearly_rev_cost:
            yearly_rev_cost[yr] = {'revenue': 0.0, 'cost': 0.0}
        yearly_rev_cost[yr]['revenue'] += amt
        yearly_rev_cost[yr]['cost'] += cost
        
        # 3. Customer Breakdown
        if cust:
            ck = (yr, cust)
            cust_breakdown[ck] = cust_breakdown.get(ck, 0.0) + qty
            
    # Format output
    volume_out = []
    for yr in sorted(yearly_vol.keys()):
        p_qty = yearly_vol[yr]['pos_qty']
        o_qty = yearly_vol[yr]['orders_qty']
        tot = p_qty + o_qty
        if p_qty > 0 and o_qty > 0:
            src = 'orders & pos'
        elif p_qty > 0:
            src = 'pos'
        else:
            src = 'orders'
        volume_out.append({
            'year': yr,
            'qty': tot,
            'source': src
        })
        
    margin_out = []
    for yr in sorted(yearly_rev_cost.keys()):
        rev = yearly_rev_cost[yr]['revenue']
        cost = yearly_rev_cost[yr]['cost']
        if rev != 0.0:
            margin_pct = round((rev - cost) / rev * 100, 1)
            margin_out.append({
                'year': yr,
                'margin_pct': margin_pct,
                'revenue_twd': rev
            })
            
    cb_out = []
    for (yr, cust), qty in sorted(cust_breakdown.items()):
        if qty > 0:
            cb_out.append({
                'year': yr,
                'customer': cust,
                'qty': qty
            })

    return jsonify({
        "query": q,
        "yearly_volume": volume_out,
        "yearly_margin": margin_out,
        "customer_breakdown": cb_out,
    })


@app.route("/api/analysis/category")
def api_analysis_category():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "請提供搜尋關鍵字"}), 400

    all_orders = db.search_orders("", limit=5000)
    all_pos = db.search_pos("", limit=20000)
    merged = get_parsed_transactions(all_orders, all_pos, q=q, filter_type='category')

    yearly_vol = {}
    yearly_rev_cost = {}
    cust_breakdown = {}
    
    for tx in merged:
        yr = tx['year']
        source = tx['source']
        qty = tx['qty']
        amt = tx['amount_twd']
        cost = tx['cost_twd']
        cust = tx['customer']
        
        # 1. Volume
        if yr not in yearly_vol:
            yearly_vol[yr] = {'pos_qty': 0.0, 'orders_qty': 0.0}
        if source == 'pos':
            yearly_vol[yr]['pos_qty'] += qty
        else:
            yearly_vol[yr]['orders_qty'] += qty
            
        # 2. Revenue & Cost (for margin)
        if yr not in yearly_rev_cost:
            yearly_rev_cost[yr] = {'revenue': 0.0, 'cost': 0.0}
        yearly_rev_cost[yr]['revenue'] += amt
        yearly_rev_cost[yr]['cost'] += cost
        
        # 3. Customer Breakdown
        if cust:
            ck = (yr, cust)
            cust_breakdown[ck] = cust_breakdown.get(ck, 0.0) + qty
            
    # Format output
    volume_out = []
    for yr in sorted(yearly_vol.keys()):
        p_qty = yearly_vol[yr]['pos_qty']
        o_qty = yearly_vol[yr]['orders_qty']
        tot = p_qty + o_qty
        if p_qty > 0 and o_qty > 0:
            src = 'orders & pos'
        elif p_qty > 0:
            src = 'pos'
        else:
            src = 'orders'
        volume_out.append({
            'year': yr,
            'qty': tot,
            'source': src
        })
        
    margin_out = []
    for yr in sorted(yearly_rev_cost.keys()):
        rev = yearly_rev_cost[yr]['revenue']
        cost = yearly_rev_cost[yr]['cost']
        if rev != 0.0:
            margin_pct = round((rev - cost) / rev * 100, 1)
            margin_out.append({
                'year': yr,
                'margin_pct': margin_pct,
                'revenue_twd': rev
            })
            
    cb_out = []
    for (yr, cust), qty in sorted(cust_breakdown.items()):
        if qty > 0:
            cb_out.append({
                'year': yr,
                'customer': cust,
                'qty': qty
            })

    return jsonify({
        "query": q,
        "yearly_volume": volume_out,
        "yearly_margin": margin_out,
        "customer_breakdown": cb_out,
    })


# ── Main ─────────────────────────────────────────────

if __name__ == "__main__":
    counts = db.get_db_counts()
    print("=" * 52)
    print("  CRM 業務與交易查詢系統 — Web 介面")
    print("=" * 52)
    if all(v == 0 for v in counts.values()):
        print("  [警告] 本地 DB 尚無資料！")
        print("  請先執行: python kintone_v2.py --sync")
    else:
        print(f"  聯絡人 : {counts['contacts']:>6,} 筆")
        print(f"  訂購單 : {counts['orders']:>6,} 筆")
        print(f"  POS    : {counts['pos_records']:>6,} 筆")
    print("  網址   : http://localhost:5000")
    print("=" * 52)
    webbrowser.open("http://localhost:5000")
    app.run(debug=False, port=5000, host="127.0.0.1")
